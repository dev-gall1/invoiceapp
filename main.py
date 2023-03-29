import pandas as pd
import openpyxl
from fpdf import FPDF

arq1 = pd.read_excel("data/10001-2023.1.18.xlsx")
arq2 = pd.read_excel("data/10002-2023.1.18.xlsx")
arq3 = pd.read_excel("data/10003-2023.1.18.xlsx")

df = pd.concat([arq1,arq2,arq3])

sumprice = df["price_per_unit"].sum()
sumtotalprice = df["total_price"].sum()

print(df["total_price"].sum())

print(f"{sumprice:.2f},{sumtotalprice:.2f}")


pdf = FPDF(orientation="L",unit="mm", format="A4")

pdf.add_page()

pdf.set_font(family="Arial",style="B",size=30)

pdf.cell(w=0,h=12,txt="INVOICE",ln=1,align="L")

pdf.set_font(family="Arial",size=10)

for index, row in df.iterrows():
    pdf.cell(w=0,h=12,txt=f"{row['product_id']}: {row['product_name']}",border=1,ln=1,align="Center")

pdf.set_font(family="Arial",size=15,style="B")
pdf.cell(w=60,h=24,txt=f'Total Unit: ${sumprice:.2f}',border=1,align="L")
pdf.cell(w=217,h=24,txt=f'Total: ${sumtotalprice:.2f}',border=1,align="R")

pdf.output("")